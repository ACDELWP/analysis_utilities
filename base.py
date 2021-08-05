from os import getenv
from os.path import isfile
from sys import exit

import sqlalchemy as sa
from openpyxl import load_workbook
from pandas import read_csv


def print_environment_variable(env_variable):
    """
    This function prints all sub-path elements in a path variable which is separated by ':'.

    :param env_variable:
    :return:
    """
    local_env_variable = str(getenv(env_variable))
    if ':' in local_env_variable:
        for subpath in str.split(local_env_variable, ':'):
            print(subpath)
    else:
        print(local_env_variable)


def return_environment_variable(env_variable):
    """
    This function returns an environment variable using os.getenv()

    :param env_variable:
    :return: env_variable as type string
    """
    try:
        return str(getenv(env_variable))
    except Exception as e:
        print(f'Tried to retrieve environment variable: {env_variable}\n')
        print(f'Failed with Exception {e}')
        return 1


def get_environment_variable(env_variable, verbose=True):
    """
    This function returns environment variable/s and has the option to also print them to device.

    USAGE:

        path = get_environment_variable('PATH')
        path, python_path = get_environment_variable(['PATH', 'PYTHONPATH'])

    :param env_variable: object of type str or list holding the desired system variables
    :param verbose: boolean which triggers verbose output to device
    :return: object of type str or list based on the input type of env_variable
    """
    if type(env_variable) not in [list, str]:
        if verbose:
            print('env_variable needs to be of type list or str.')
        return 1
    if type(env_variable) is str:
        if verbose:
            print_environment_variable(env_variable)
        return return_environment_variable(env_variable)
    if type(env_variable) is list:
        return_object = []
        for local_env_variable in env_variable:
            try:
                if type(local_env_variable) is str:
                    if verbose:
                        print_environment_variable(local_env_variable)
                    return_object.append(return_environment_variable(local_env_variable))
            except Exception as e:
                print(f'Tried to find ENVIRONMENT VARIABLE: {local_env_variable}\n')
                print(f'Failed with Exception: {e}')
                return_object.append(None)

        return return_object


def load_user_profile(user=None, db=None):
    current_accepted_users_and_location = {
        'acodoreanu_local': r'C:\Users\ac4o\OneDrive - Department of Environment, Land, Water and Planning\Projects\GitHub\Mer-data\core\dev'}
    if user is not None:
        if db is not None:
            if user in list(current_accepted_users_and_location.keys()):
                if len(list(current_accepted_users_and_location.keys())) > 1:
                    print(f'{user} has multiple credential locations')
                    return 1

                local_user = list(current_accepted_users_and_location.keys())[0]
                location = current_accepted_users_and_location[local_user]
                credential_file_name = f'{location}/{user}_{db}_credentials.csv'
                if isfile(credential_file_name):
                    try:
                        return read_csv(credential_file_name, sep=';')
                    except Exception as e:
                        print(f'Could not open the following credentials file:\n{credential_file_name}')
                        print(f'\nFailed with exception:\n{e}')
                        return 1
                else:
                    print(f'The following credentials file does not exist:\n{credential_file_name}')
                    return 1
            else:
                print(f'{user} is not a recognised key in current_accepted_users_and_location dictionary')
                return 1
        else:
            print(f'You did not pass me a database to connect to for user:{user}')
            return 1
    else:
        print('You did not pass me a user name.')
        return 1


def get_sql_alchemy_engine(host=None, db=None, user_name=None, user_password=None,
                           driver="ODBC Driver 17 for SQL Server"):
    if any(elem is None for elem in [host, db, user_name, user_password]):
        print("You must pass a value for all of the following variables: [host, db, user_name, user_password]\n")
        return 1

    connection_url = sa.engine.url.URL.create(
        "mssql+pyodbc",
        username=user_name,
        password=user_password,
        host=host,
        database=db,
        query={"driver": driver},
    )

    engine = sa.create_engine(connection_url, fast_executemany=True)

    return engine


def get_sheetnames_excel_file(filename=None):
    """
    This function returns all sheets inide of an excel file.

    :param filename: excel file with full path
    :return: sheets (or exits with error message)
    """
    if filename is None:
        exit('You did not pass me an excel file name to open.\n')

    try:
        return load_workbook(filename, read_only=True).sheetnames
    except Exception as e:
        exit(e)
